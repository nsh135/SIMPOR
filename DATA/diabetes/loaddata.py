import numpy as np
# Import openyxl module
import openpyxl
import os
# def load_diabetes():
#     """
#     Load diabetes data
#     return numpy X,y {0: normal, 1: diabetes}
#     """
#     dir = os.path.dirname (os.path.abspath(__file__))
#     # Give the location of the file
#     path = (dir + "/diabetes.xlsx")
#     # Define variable to load the wookbook
#     wookbook = openpyxl.load_workbook(path)
#     # Define variable to read the active sheet:
#     worksheet = wookbook.active
#     r = 467
#     c = 9
#     M = [ [ 0 for i in range(c) ] for j in range(r) ]

#     # Iterate the loop to read the cell values
#     for i in range(1, worksheet.max_row):
#         j=0
#         for col in worksheet.iter_cols(1, worksheet.max_column):
# #             print(col[i].value, end="\t")
#             M[i-1][j] = col[i].value
#             j +=1
# #         print('')
#     M=np.array(M)
#     X = M[:,1:8]
#     y = np.array([0 if label ==1 else 1 for label in M[:,0]] )
#     return X,y



#### Pima Indians Diabetes dataset
from numpy import unique
from pandas import read_csv
# load the dataset
def load_diabetes():
	url = 'https://raw.githubusercontent.com/jbrownlee/Datasets/master/pima-indians-diabetes.csv'
	dataframe = read_csv(url, header=None)
	# get the values
	values = dataframe.values
	X, y = values[:, :-1], values[:, -1]
	# gather details
	n_rows = X.shape[0]
	n_cols = X.shape[1]
	classes = unique(y)
	n_classes = len(classes)
	# summarize
	print('N Examples: %d' % n_rows)
	print('N Inputs: %d' % n_cols)
	print('N Classes: %d' % n_classes)
	print('Classes: %s' % classes)
	print('Class Breakdown:')
	# class breakdown
	breakdown = ''
	for c in classes:
		total = len(y[y == c])
		ratio = (total / float(len(y))) * 100
		print(' - Class %s: %d (%.5f%%)' % (str(c), total, ratio))
	return X,y